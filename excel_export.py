"""Excel export utilities for multi-sheet KPI reports"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
from io import BytesIO
import os

# Try to import image support
try:
    from openpyxl.drawing.image import Image as ExcelImage
    IMAGE_SUPPORT = True
except ImportError:
    IMAGE_SUPPORT = False
    print("Warning: Image support not available in openpyxl")

def create_excel_report(sheets_data: Dict[str, List[List]]) -> BytesIO:
    """
    Create a multi-sheet Excel workbook from sheets_data dict.
    
    Args:
        sheets_data: Dict where keys are sheet names and values are 2D lists of data
        
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Style definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Check if logo exists
    logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo_gsit_ss.png')
    has_logo = IMAGE_SUPPORT and os.path.exists(logo_path)
    
    if not IMAGE_SUPPORT:
        print("Warning: PIL/Pillow not installed - images disabled")
    if not os.path.exists(logo_path):
        print(f"Warning: Logo file not found at {logo_path}")
    
    # Create sheets
    first_sheet_processed = False
    for sheet_idx, (sheet_name, data) in enumerate(sheets_data.items()):
        ws = wb.create_sheet(title=sheet_name)

        if isinstance(data, dict) and "rows" in data:
            sheet_rows = data.get("rows", [])
            sheet_groups = data.get("groups", [])
        else:
            sheet_rows = data
            sheet_groups = []
        
        # Add logo to first sheet and Executive Summary/Summary sheets
        add_logo = (not first_sheet_processed) or ('Summary' in sheet_name or 'SUMMARY' in sheet_name)
        
        if add_logo and has_logo:
            try:
                print(f"Adding logo to sheet: {sheet_name}")
                img = ExcelImage(logo_path)
                # Resize logo to fit nicely
                img.width = 120
                img.height = 60
                # Place at E1 (top-right area, won't interfere with main content)
                ws.add_image(img, 'E1')
                print(f"Logo added successfully to {sheet_name}")
                if not first_sheet_processed:
                    first_sheet_processed = True
            except Exception as e:
                print(f"Error adding logo to {sheet_name}: {e}")
                import traceback
                traceback.print_exc()
        
        # Write data
        start_row = 1
        for row_idx, row_data in enumerate(sheet_rows, start_row):
            is_date_device_row = False
            first_cell = row_data[0] if row_data else None
            if isinstance(first_cell, str) and (first_cell.startswith('DATE:') or first_cell.startswith('DEVICE:')):
                is_date_device_row = True

            for col_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                cell.border = border
                
                # Style first row (title row) specially if it contains report title
                if row_idx == 1 and isinstance(cell_value, str) and 'REPORT' in cell_value.upper():
                    cell.font = Font(bold=True, size=16, color="1F4E78")
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                # Style headers
                elif row_idx > 1 and any(isinstance(cell_value, str) and cell_value in ['Key Metric', 'Metric', 'Rank', 'Engineer'] for cell_value in row_data):
                    cell.fill = header_fill
                    cell.font = header_font
                # Style section headers (rows with single merged cell or key column)
                elif isinstance(cell_value, str) and cell_value.isupper() and len(str(cell_value)) > 3:
                    # Check if it's a section header (all caps, longer than 3 chars)
                    if col_idx == 1 and any(cell_value.startswith(x) for x in ['EXECUTIVE', 'PERFORMANCE', 'TARGET', 'TOP', 'ALL', 'BREAKDOWN', 'ENGINEER', 'HISTORICAL', 'WEEKLY', 'SPEED', 'CATEGORY', 'CONSISTENCY', 'REPORT', 'GLOSSARY', 'DATE', 'DEVICE']):
                        cell.fill = section_fill
                        cell.font = section_font
                if is_date_device_row:
                    cell.fill = section_fill
                    cell.font = section_font
        
        # Apply row grouping for collapsible sections
        for group in sheet_groups:
            if len(group) >= 4:
                start, end, level, hidden = group
                ws.row_dimensions.group(start, end, outline_level=level, hidden=hidden)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
